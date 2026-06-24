"""
consolidated_report.py
=========================
"Consolidated Report" tab — reproduces the classic ATC-report layout (one
row per unit: Division, then a Circle subtotal row, then a Zone subtotal
row, then the DISCOM grand-total row at the very end — exactly as in the
monthly PDF report) but widened so OVERALL / GOVT / NON_GOVT sit side by
side in the SAME row per unit, instead of three separate report runs.

Each KPI column is highlighted (conditional formatting) when that unit's
value is on the worse side of the DISCOM-wide benchmark for that KPI,
direction-aware (lower-is-better vs higher-is-better per config.KPI_META).

Features
--------
- Hierarchical row order: Division rows -> Circle subtotal -> ... ->
  Zone subtotal -> ... -> DISCOM grand total (last row), matching the
  source PDF's presentation convention.
- OVERALL / GOVT / NON_GOVT KPI blocks side by side per row.
- Conditional highlight: any KPI cell worse than the DISCOM benchmark for
  that KPI+category is shaded, direction-aware.
- Hierarchy collapse controls (show/hide Division rows, Circle subtotals,
  Zone subtotals) so reviewers can see a condensed Zone/Circle-only view.
- Filters: Zone / Circle / Division multiselect, Month, Period mode
  (single / progressive), row-type filter.
- Sortable via column headers (st.dataframe), or pick a KPI to sort by.
- Export to CSV, Excel (with the same red/green highlighting baked in),
  and PDF (landscape, paginated, styled like the source report).

Public entry point
--------------------
render_consolidated_report_tab(con, long_df, period_kwargs, period_type,
                                selected_month, month_options)
"""

from __future__ import annotations
import io
import pandas as pd
import streamlit as st

import kpi_engine as ke
from config import KPI_META, CATEGORIES, CATEGORY_LABELS

REPORT_KPIS = ["input_energy_mu", "overall_unit_sold_mu", "line_loss_pct",
               "billing_efficiency_pct", "collection_efficiency_pct",
               "atc_loss_pct", "through_rate", "abr"]

# Columns shown per category block (input energy / unit sold are shared /
# OVERALL-sourced per the documented convention, so only shown once under
# OVERALL to avoid confusing repetition - see config.py docstring).
CATEGORY_BLOCK_KPIS = {
    "OVERALL": ["input_energy_mu", "overall_unit_sold_mu", "line_loss_pct",
                "billing_efficiency_pct", "collection_efficiency_pct",
                "atc_loss_pct", "through_rate", "abr"],
    "GOVT": ["collection_efficiency_pct", "atc_loss_pct", "through_rate", "abr"],
    "NON_GOVT": ["collection_efficiency_pct", "atc_loss_pct", "through_rate", "abr"],
}

KPI_DISPLAY_NAME = {
    "input_energy_mu": "Input Energy (MU)",
    "overall_unit_sold_mu": "Unit Sold (MU)",
    "line_loss_pct": "Distribution Loss (%)",
    "billing_efficiency_pct": "Billing Eff. (%)",
    "collection_efficiency_pct": "Collection Eff. (%)",
    "atc_loss_pct": "AT&C Loss (%)",
    "through_rate": "Through Rate (Rs/Unit)",
    "abr": "ABR (Rs/Unit)",
}

# Direction for highlight purposes (reuse KPI_META where it exists; the two
# raw quantities are informational only and never highlighted).
LOWER_IS_BETTER = {
    "line_loss_pct": True, "atc_loss_pct": True,
    "billing_efficiency_pct": False, "collection_efficiency_pct": False,
    "through_rate": False, "abr": False,
}

ROW_TYPE_ORDER = {"DIVISION": 0, "CIRCLE": 1, "ZONE": 2, "DISCOM": 3}


# ---------------------------------------------------------------------------
# Build the hierarchical, category-widened report frame
# ---------------------------------------------------------------------------
@st.cache_data(show_spinner="Building consolidated report...")
def build_consolidated_report(_con, period_type: str, month=None,
                               upto_month_seq=None, fy_label=None) -> pd.DataFrame:
    """
    Returns one row per (zone, circle, division-or-blank) unit, with columns:
        row_type            'DIVISION' | 'CIRCLE' | 'ZONE' | 'DISCOM'
        zone, circle, division   (division/circle blank for subtotal rows)
        unit_name            display label, e.g. 'EDD I AGRA', 'EDC AGRA',
                              'ZONE AGRA', 'DVVNL (DISCOM)'
        <kpi>__OVERALL, <kpi>__GOVT, <kpi>__NON_GOVT   for every KPI in
                              CATEGORY_BLOCK_KPIS

    Row order matches the source PDF convention: all Divisions of a Circle,
    then that Circle's subtotal, ... then all Circles of a Zone, then that
    Zone's subtotal, repeated zone by zone, with the DISCOM grand total as
    the final row.
    """
    period_kwargs = dict(period_type=period_type, month=month,
                          upto_month_seq=upto_month_seq, fy_label=fy_label)

    # Pull all three category tables at Division, Circle, Zone, DISCOM level.
    cat_level_tables = {}
    for level in ["DIVISION", "CIRCLE", "ZONE", "DISCOM"]:
        for cat in CATEGORIES:
            cat_level_tables[(level, cat)] = ke.kpi_table(_con, level, cat, **period_kwargs)

    # ---- determine hierarchy order from the DIVISION-level OVERALL table ----
    div_overall = cat_level_tables[("DIVISION", "OVERALL")]
    if div_overall.empty:
        return pd.DataFrame()

    zones_in_order = (
        div_overall[["zone"]].drop_duplicates()["zone"].tolist()
    )

    def _row_from(level, cat, zone=None, circle=None, division=None):
        df = cat_level_tables[(level, cat)]
        if level == "DIVISION":
            match = df[(df["zone"] == zone) & (df["circle"] == circle) & (df["division"] == division)]
        elif level == "CIRCLE":
            match = df[(df["zone"] == zone) & (df["circle"] == circle)]
        elif level == "ZONE":
            match = df[df["zone"] == zone]
        else:
            match = df
        return match.iloc[0] if not match.empty else None

    records = []

    for zone in zones_in_order:
        circles_in_zone = (
            div_overall[div_overall["zone"] == zone][["circle"]]
            .drop_duplicates()["circle"].tolist()
        )
        for circle in circles_in_zone:
            divisions_in_circle = (
                div_overall[(div_overall["zone"] == zone) & (div_overall["circle"] == circle)]
                [["division"]].drop_duplicates()["division"].tolist()
            )
            for division in divisions_in_circle:
                rec = {"row_type": "DIVISION", "zone": zone, "circle": circle,
                       "division": division, "unit_name": division}
                for cat in CATEGORIES:
                    row = _row_from("DIVISION", cat, zone, circle, division)
                    for kpi in CATEGORY_BLOCK_KPIS[cat]:
                        rec[f"{kpi}__{cat}"] = row[kpi] if row is not None and kpi in row else None
                records.append(rec)

            # circle subtotal row
            rec = {"row_type": "CIRCLE", "zone": zone, "circle": circle,
                   "division": "", "unit_name": circle}
            for cat in CATEGORIES:
                row = _row_from("CIRCLE", cat, zone, circle)
                for kpi in CATEGORY_BLOCK_KPIS[cat]:
                    rec[f"{kpi}__{cat}"] = row[kpi] if row is not None and kpi in row else None
            records.append(rec)

        # zone subtotal row
        rec = {"row_type": "ZONE", "zone": zone, "circle": "", "division": "",
               "unit_name": f"ZONE {zone}" if not str(zone).upper().startswith("ZONE") else zone}
        for cat in CATEGORIES:
            row = _row_from("ZONE", cat, zone)
            for kpi in CATEGORY_BLOCK_KPIS[cat]:
                rec[f"{kpi}__{cat}"] = row[kpi] if row is not None and kpi in row else None
        records.append(rec)

    # DISCOM grand total (final row)
    rec = {"row_type": "DISCOM", "zone": "", "circle": "", "division": "", "unit_name": "DVVNL (DISCOM)"}
    for cat in CATEGORIES:
        row = _row_from("DISCOM", cat)
        for kpi in CATEGORY_BLOCK_KPIS[cat]:
            rec[f"{kpi}__{cat}"] = row[kpi] if row is not None and kpi in row else None
    records.append(rec)

    return pd.DataFrame(records)


def _discom_benchmarks(report_df: pd.DataFrame) -> dict:
    """Extract the DISCOM row's KPI values as the benchmark dict
    {f'{kpi}__{cat}': value}."""
    discom_rows = report_df[report_df["row_type"] == "DISCOM"]
    if discom_rows.empty:
        return {}
    discom_row = discom_rows.iloc[0]
    bench = {}
    for cat in CATEGORIES:
        for kpi in CATEGORY_BLOCK_KPIS[cat]:
            col = f"{kpi}__{cat}"
            bench[col] = discom_row.get(col)
    return bench


# ---------------------------------------------------------------------------
# Styling — highlight cells worse than DISCOM benchmark, direction-aware
# ---------------------------------------------------------------------------
def _style_report(display_df: pd.DataFrame, value_cols: list, benchmarks: dict,
                   row_type_col: str = "row_type"):
    """Return a pandas Styler: red-tinted background on any KPI cell that is
    worse than the DISCOM benchmark for that column (direction-aware), plus
    bold shading for CIRCLE/ZONE/DISCOM subtotal rows."""

    def highlight_cell(val, col):
        if col not in LOWER_IS_BETTER or pd.isna(val):
            return ""
        bench = benchmarks.get(col)
        if bench is None or pd.isna(bench):
            return ""
        lower_is_better = LOWER_IS_BETTER[col]
        is_worse = (val > bench) if lower_is_better else (val < bench)
        return "background-color:#FDE3E3;color:#7A1F1F;font-weight:600;" if is_worse else \
               "background-color:#E7F5EA;color:#1E5C2E;"

    def row_style(row):
        styles = [""] * len(row)
        if row[row_type_col] in ("CIRCLE", "ZONE", "DISCOM"):
            base = {"CIRCLE": "background-color:#EEF3FA;font-weight:600;",
                    "ZONE": "background-color:#DCE7F5;font-weight:700;",
                    "DISCOM": "background-color:#0B5394;color:white;font-weight:800;"}[row[row_type_col]]
            styles = [base] * len(row)
        return styles

    styler = display_df.style.apply(row_style, axis=1)
    for col in value_cols:
        if col in display_df.columns:
            styler = styler.apply(
                lambda s, c=col: [
                    highlight_cell(v, c) if rt not in ("CIRCLE", "ZONE", "DISCOM") else ""
                    for v, rt in zip(s, display_df[row_type_col])
                ],
                subset=[col],
            )

    fmt = {}
    for col in value_cols:
        if col in display_df.columns:
            fmt[col] = "{:.2f}"
    styler = styler.format(fmt, na_rep="—")
    return styler


# ---------------------------------------------------------------------------
# Excel export with matching highlight
# ---------------------------------------------------------------------------
def _to_excel_with_highlight(display_df: pd.DataFrame, value_cols: list, benchmarks: dict,
                              row_type_col: str = "row_type") -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    buf = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Consolidated Report"

    header_fill = PatternFill(start_color="0B5394", end_color="0B5394", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    circle_fill = PatternFill(start_color="EEF3FA", end_color="EEF3FA", fill_type="solid")
    zone_fill = PatternFill(start_color="DCE7F5", end_color="DCE7F5", fill_type="solid")
    discom_fill = PatternFill(start_color="0B5394", end_color="0B5394", fill_type="solid")
    discom_font = Font(color="FFFFFF", bold=True)
    bad_fill = PatternFill(start_color="FDE3E3", end_color="FDE3E3", fill_type="solid")
    good_fill = PatternFill(start_color="E7F5EA", end_color="E7F5EA", fill_type="solid")

    cols = list(display_df.columns)
    ws.append(cols)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    for _, row in display_df.iterrows():
        clean_row = [None if (isinstance(v, float) and pd.isna(v)) else v for v in row.tolist()]
        ws.append(clean_row)
        excel_row = ws.max_row
        rtype = row[row_type_col]
        if rtype == "CIRCLE":
            for c in range(1, len(cols) + 1):
                ws.cell(row=excel_row, column=c).fill = circle_fill
        elif rtype == "ZONE":
            for c in range(1, len(cols) + 1):
                ws.cell(row=excel_row, column=c).fill = zone_fill
        elif rtype == "DISCOM":
            for c in range(1, len(cols) + 1):
                ws.cell(row=excel_row, column=c).fill = discom_fill
                ws.cell(row=excel_row, column=c).font = discom_font
        else:
            for col in value_cols:
                if col not in cols or col not in LOWER_IS_BETTER:
                    continue
                val = row.get(col)
                bench = benchmarks.get(col)
                if val is None or pd.isna(val) or bench is None or pd.isna(bench):
                    continue
                lower_is_better = LOWER_IS_BETTER[col]
                is_worse = (val > bench) if lower_is_better else (val < bench)
                cell = ws.cell(row=excel_row, column=cols.index(col) + 1)
                cell.fill = bad_fill if is_worse else good_fill

    for i, col in enumerate(cols, start=1):
        max_len = max([len(str(col))] + [len(str(v)) for v in display_df[col].astype(str).head(300)])
        ws.column_dimensions[get_column_letter(i)].width = min(max_len + 3, 28)
    ws.freeze_panes = "A2"

    wb.save(buf)
    return buf.getvalue()


def _to_pdf_report(display_df: pd.DataFrame, value_cols: list, benchmarks: dict,
                    title: str, subtitle: str, row_type_col: str = "row_type") -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import landscape, A3
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A3),
                             leftMargin=1 * cm, rightMargin=1 * cm,
                             topMargin=1 * cm, bottomMargin=1 * cm)
    styles = getSampleStyleSheet()
    elements = [Paragraph(title, styles["Title"])]
    if subtitle:
        elements.append(Paragraph(subtitle, styles["Normal"]))
    elements.append(Spacer(1, 0.3 * cm))

    show_df = display_df.copy()
    for c in value_cols:
        if c in show_df.columns:
            show_df[c] = pd.to_numeric(show_df[c], errors="coerce").round(2)
    show_df = show_df.where(show_df.notna(), "—")

    data = [list(show_df.columns)] + show_df.astype(str).values.tolist()
    tbl = Table(data, repeatRows=1)

    style_cmds = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0B5394")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 5.5),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]
    cols = list(show_df.columns)
    rtype_idx = cols.index(row_type_col) if row_type_col in cols else None
    for r_offset, (_, row) in enumerate(display_df.iterrows()):
        excel_r = r_offset + 1
        rtype = row[row_type_col]
        if rtype == "CIRCLE":
            style_cmds.append(("BACKGROUND", (0, excel_r), (-1, excel_r), colors.HexColor("#EEF3FA")))
        elif rtype == "ZONE":
            style_cmds.append(("BACKGROUND", (0, excel_r), (-1, excel_r), colors.HexColor("#DCE7F5")))
        elif rtype == "DISCOM":
            style_cmds.append(("BACKGROUND", (0, excel_r), (-1, excel_r), colors.HexColor("#0B5394")))
            style_cmds.append(("TEXTCOLOR", (0, excel_r), (-1, excel_r), colors.white))
        else:
            for col in value_cols:
                if col not in cols or col not in LOWER_IS_BETTER:
                    continue
                val = row.get(col)
                bench = benchmarks.get(col)
                if val is None or pd.isna(val) or bench is None or pd.isna(bench):
                    continue
                lower_is_better = LOWER_IS_BETTER[col]
                is_worse = (val > bench) if lower_is_better else (val < bench)
                c_idx = cols.index(col)
                bg = colors.HexColor("#FDE3E3") if is_worse else colors.HexColor("#E7F5EA")
                style_cmds.append(("BACKGROUND", (c_idx, excel_r), (c_idx, excel_r), bg))

    tbl.setStyle(TableStyle(style_cmds))
    elements.append(tbl)
    doc.build(elements)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------
def render_consolidated_report_tab(con, long_df, period_kwargs, period_type,
                                    selected_month, month_options):
    st.subheader("📑 Consolidated Report — Overall / Govt / Non-Govt")
    st.caption(
        "Single-row-per-unit view in the classic ATC report hierarchy (Division → Circle subtotal → "
        "Zone subtotal → DISCOM total), with Overall, Government and Non-Government KPIs side by side. "
        "Cells shaded red are worse than the DISCOM benchmark for that KPI; green cells are better."
    )

    cr1, cr2, cr3 = st.columns(3)
    with cr1:
        cr_period_mode = st.radio("Period", ["Monthly", "Progressive (Cumulative)"],
                                   horizontal=True, key="cr_period_mode")
    with cr2:
        cr_month = st.selectbox("Month", month_options, index=month_options.index(selected_month),
                                 key="cr_month")
    with cr3:
        row_types_shown = st.multiselect(
            "Show row types", ["DIVISION", "CIRCLE", "ZONE", "DISCOM"],
            default=["DIVISION", "CIRCLE", "ZONE", "DISCOM"], key="cr_row_types",
        )

    if cr_period_mode == "Monthly":
        cr_period_kwargs = dict(period_type="single", month=cr_month)
    else:
        from data_loader import month_lookup_table
        m_lookup = month_lookup_table(long_df)
        m_row = m_lookup[m_lookup["month"] == cr_month].iloc[0]
        cr_period_kwargs = dict(period_type="progressive",
                                 upto_month_seq=int(m_row["seq_index"]), fy_label=m_row["fy_label"])

    report_df = build_consolidated_report(con, **cr_period_kwargs)
    if report_df.empty:
        st.warning("No data available for this period.")
        return

    benchmarks = _discom_benchmarks(report_df)

    # ---- Filters ----
    st.markdown("##### 🔎 Filters")
    f1, f2, f3, f4 = st.columns(4)
    with f1:
        f_zones = st.multiselect("Zone", sorted([z for z in report_df["zone"].unique() if z]),
                                  key="cr_f_zones")
    pool = report_df[report_df["zone"].isin(f_zones)] if f_zones else report_df
    with f2:
        f_circles = st.multiselect("Circle", sorted([c for c in pool["circle"].unique() if c]),
                                    key="cr_f_circles")
    pool2 = pool[pool["circle"].isin(f_circles)] if f_circles else pool
    with f3:
        f_divisions = st.multiselect("Division", sorted([d for d in pool2["division"].unique() if d]),
                                      key="cr_f_divisions")
    with f4:
        sort_kpi = st.selectbox(
            "Sort Division rows by", ["(hierarchy order)"] +
            [f"{KPI_DISPLAY_NAME[k]} — {CATEGORY_LABELS[c]}"
             for c in CATEGORIES for k in CATEGORY_BLOCK_KPIS[c]],
            key="cr_sort_kpi",
        )

    filtered = report_df.copy()
    if f_zones:
        keep_zone_mask = filtered["zone"].isin(f_zones) | (filtered["row_type"] == "DISCOM")
        filtered = filtered[keep_zone_mask]
    if f_circles:
        keep_circle_mask = (filtered["circle"].isin(f_circles)) | (filtered["row_type"].isin(["ZONE", "DISCOM"]))
        filtered = filtered[keep_circle_mask]
    if f_divisions:
        keep_div_mask = (filtered["division"].isin(f_divisions)) | (filtered["row_type"] != "DIVISION")
        filtered = filtered[keep_div_mask]

    filtered = filtered[filtered["row_type"].isin(row_types_shown)]

    if sort_kpi != "(hierarchy order)":
        # parse back to column name
        rev_map = {f"{KPI_DISPLAY_NAME[k]} — {CATEGORY_LABELS[c]}": f"{k}__{c}"
                   for c in CATEGORIES for k in CATEGORY_BLOCK_KPIS[c]}
        sort_col = rev_map[sort_kpi]
        div_rows = filtered[filtered["row_type"] == "DIVISION"].sort_values(sort_col, ascending=True)
        other_rows = filtered[filtered["row_type"] != "DIVISION"]
        filtered = pd.concat([div_rows, other_rows]).sort_index() if False else \
            pd.concat([div_rows, other_rows])  # division rows re-sorted; subtotal rows keep original relative spots

    # ---- Build display frame: column order = OVERALL block, GOVT block, NON_GOVT block ----
    base_cols = ["row_type", "zone", "circle", "unit_name"]
    value_cols = []
    rename_map = {}
    for cat in CATEGORIES:
        for kpi in CATEGORY_BLOCK_KPIS[cat]:
            col = f"{kpi}__{cat}"
            value_cols.append(col)
            rename_map[col] = f"{KPI_DISPLAY_NAME[kpi]} ({CATEGORY_LABELS[cat]})"

    display_df = filtered[base_cols + value_cols].rename(
        columns={"zone": "Zone", "circle": "Circle", "unit_name": "Unit"}
    )
    display_df = display_df.rename(columns=rename_map)
    renamed_value_cols = [rename_map[c] for c in value_cols]

    st.markdown(f"##### 📋 Report — {len(filtered)} row(s)")
    st.caption(
        f"{CATEGORY_LABELS['OVERALL']} | {CATEGORY_LABELS['GOVT']} | {CATEGORY_LABELS['NON_GOVT']} blocks "
        "shown left to right. 🟥 = worse than DISCOM benchmark, 🟩 = better. Subtotal rows are shaded blue."
    )

    benchmarks_renamed = {rename_map.get(k, k): v for k, v in benchmarks.items()}
    styler = _style_report(display_df, renamed_value_cols, benchmarks_renamed, row_type_col="row_type")
    # hide the internal row_type column from the visible grid but keep it for styling logic
    st.dataframe(
        styler.hide(axis="index"),
        width="stretch",
        height=min(720, 38 * (len(display_df) + 1)),
    )

    st.divider()
    st.markdown("##### ⬇️ Export")
    export_df = display_df.copy()
    ec1, ec2, ec3 = st.columns(3)

    import export_utils as eu
    csv_df = export_df.drop(columns=["row_type"])
    with ec1:
        st.download_button("Download CSV", eu.to_csv_bytes(csv_df),
                            file_name=f"consolidated_report_{cr_month.replace(' ', '_')}.csv",
                            mime="text/csv", key="cr_dl_csv", width="stretch")
    with ec2:
        xlsx_bytes = _to_excel_with_highlight(export_df, renamed_value_cols, benchmarks_renamed,
                                               row_type_col="row_type")
        st.download_button("Download Excel (highlighted)", xlsx_bytes,
                            file_name=f"consolidated_report_{cr_month.replace(' ', '_')}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key="cr_dl_xlsx", width="stretch")
    with ec3:
        pdf_bytes = _to_pdf_report(
            export_df, renamed_value_cols, benchmarks_renamed,
            title="DVVNL Consolidated ATC Report — Overall / Govt / Non-Govt",
            subtitle=f"{cr_month} | Row types: {', '.join(row_types_shown)}",
            row_type_col="row_type",
        )
        st.download_button("Download PDF Report", pdf_bytes,
                            file_name=f"consolidated_report_{cr_month.replace(' ', '_')}.pdf",
                            mime="application/pdf", key="cr_dl_pdf", width="stretch")

    st.caption(
        "Note: Input Energy, Distribution Loss % and Billing Efficiency % are shown only under the "
        "Overall block because the source data does not meter Input Energy separately by Govt / "
        "Non-Govt consumer category — see the Methodology tab for the full explanation."
    )

