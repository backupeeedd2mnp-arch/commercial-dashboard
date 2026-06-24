"""
export_utils.py
================
Export helpers for the dashboard's "Data Explorer & Export" tab.

  to_excel_bytes(df, summary_title)  -> formatted .xlsx (openpyxl) as bytes
  to_csv_bytes(df)                   -> .csv as bytes
  to_pdf_bytes(df, title, chart_fig) -> one-page PDF report (table + optional
                                         chart image) using reportlab
"""

from __future__ import annotations
import io
import pandas as pd


def to_csv_bytes(df: pd.DataFrame) -> bytes:
    return df.to_csv(index=False).encode("utf-8-sig")


def to_excel_bytes(df: pd.DataFrame, sheet_title: str = "Data") -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    buf = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title[:31] if sheet_title else "Data"

    # openpyxl cannot write pandas' pd.NA / NaT directly -> normalize to None first
    df = df.where(df.notna(), None)

    header_fill = PatternFill(start_color="0B5394", end_color="0B5394", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    ws.append(list(df.columns))
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row in df.itertuples(index=False):
        clean_row = [None if (isinstance(v, float) and pd.isna(v)) or v is pd.NA else v for v in row]
        ws.append(clean_row)

    for i, col in enumerate(df.columns, start=1):
        max_len = max([len(str(col))] + [len(str(v)) for v in df[col].astype(str).head(200)])
        ws.column_dimensions[get_column_letter(i)].width = min(max_len + 3, 40)

    ws.freeze_panes = "A2"
    wb.save(buf)
    return buf.getvalue()


def to_pdf_bytes(df: pd.DataFrame, title: str = "DVVNL Management Review Report",
                  subtitle: str = "", max_rows: int = 40) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import landscape, A4
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                             leftMargin=1.2 * cm, rightMargin=1.2 * cm,
                             topMargin=1.2 * cm, bottomMargin=1.2 * cm)
    styles = getSampleStyleSheet()
    elements = [Paragraph(title, styles["Title"])]
    if subtitle:
        elements.append(Paragraph(subtitle, styles["Normal"]))
    elements.append(Spacer(1, 0.4 * cm))

    show_df = df.head(max_rows).copy()
    for c in show_df.select_dtypes(include="float").columns:
        show_df[c] = show_df[c].round(2)
    show_df = show_df.where(show_df.notna(), "N/A")

    data = [list(show_df.columns)] + show_df.astype(str).values.tolist()
    tbl = Table(data, repeatRows=1)
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0B5394")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 6.5),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F6FB")]),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))
    elements.append(tbl)

    if len(df) > max_rows:
        elements.append(Spacer(1, 0.3 * cm))
        elements.append(Paragraph(
            f"Showing first {max_rows} of {len(df)} rows. Export to Excel/CSV for the full dataset.",
            styles["Italic"]))

    doc.build(elements)
    return buf.getvalue()
